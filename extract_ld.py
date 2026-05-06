"""
extract_ld.py — Extract tag number dari PDF Logic Diagram (Safety Manager)
               dan buat file Excel mapping: Sheet → Tag Number

Cara pakai:
    python extract_ld.py

Output:
    data/LD_mapping.xlsx  ← buka dan tambahkan link GDrive per sheet
    
Setelah LD_mapping.xlsx diisi link → jalankan:
    python update_ld.py

Struktur PDF yang dibaca:
    - Page 1-2  : Cover & TOC (dilewati)
    - Page 3-5  : Cross-reference index (Tag → FLD/Sheet number)
                  Format: Type | Tag number | FLD
    - Page 6+   : Diagram per sheet
                  Format: F <TagNumber> ... di kolom kiri
                  Footer : SM_NH3  <sheet_number>  <next_sheet>

Strategi:
    Gunakan halaman cross-reference (page 3-5) yang sudah berisi
    mapping lengkap Tag → FLD (sheet number). Ini LEBIH AKURAT
    daripada extract dari diagram karena sudah terstruktur.
"""

import re
import os
import sys

# ─────────────────────────────────────────────
# KONFIGURASI
# ─────────────────────────────────────────────
PDF_PATH   = "data/SM_NH3_19_MEI_2025.pdf"
OUTPUT_XLS = "data/LD_mapping.xlsx"

# Halaman cross-reference index di PDF (page 3 s/d 6, index mulai 1)
XREF_START_PAGE = 3
XREF_END_PAGE   = 6

# ─────────────────────────────────────────────
# STEP 1: EXTRACT CROSS-REFERENCE INDEX
# ─────────────────────────────────────────────
def extract_xref(pdf_path: str) -> dict:
    """
    Baca halaman cross-reference dan return:
    {
        sheet_number: [tag1, tag2, ...],
        ...
    }
    Format baris: Type  Tag_number  FLD
    Contoh      : DI    TA1200      11
    """
    try:
        import pdfplumber
    except ImportError:
        print("❌ pdfplumber tidak terinstall. Jalankan: pip install pdfplumber")
        sys.exit(1)

    sheet_to_tags = {}  # {sheet_num: [tag, ...]}

    # Pola tag yang valid — huruf diikuti angka, bisa ada strip/titik
    TAG_PATTERN = re.compile(
        r'^[A-Z]{1,6}[\-_]?[A-Z0-9][\w\-\.]*\d+[A-Z]?$'
    )

    # Pola baris cross-reference: Type  TagNumber  FLD
    LINE_PATTERN = re.compile(
        r'(DI|DO|AI|AO|BI|BO)\s+(\S+)\s+(\d+)'
    )

    print(f"📖 Membaca cross-reference index (halaman {XREF_START_PAGE}-{XREF_END_PAGE})...")

    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        end_page = min(XREF_END_PAGE, total_pages)

        for page_num in range(XREF_START_PAGE - 1, end_page):
            page = pdf.pages[page_num]
            text = page.extract_text()
            if not text:
                continue

            for line in text.split('\n'):
                matches = LINE_PATTERN.findall(line)
                for match in matches:
                    sig_type, tag, fld = match
                    fld_num = int(fld)
                    tag = tag.strip()

                    # Filter tag yang valid
                    if not tag or len(tag) < 3:
                        continue

                    # Skip tag yang terlalu generic (fungsi internal SM)
                    skip_prefixes = ['FL0', 'FL1', 'FL2', 'SYS_', '1-', 'Digital_']
                    if any(tag.startswith(p) for p in skip_prefixes):
                        continue

                    if fld_num not in sheet_to_tags:
                        sheet_to_tags[fld_num] = []
                    if tag not in sheet_to_tags[fld_num]:
                        sheet_to_tags[fld_num].append(tag)

    total_tags = sum(len(v) for v in sheet_to_tags.values())
    print(f"✅ Ditemukan {len(sheet_to_tags)} sheet, {total_tags} tag total")
    return sheet_to_tags


# ─────────────────────────────────────────────
# STEP 2: BUAT EXCEL MAPPING
# ─────────────────────────────────────────────
def create_excel(sheet_to_tags: dict, output_path: str):
    """
    Buat Excel dengan format:
    Baris 1: Page    | Sheet 11 | Sheet 13 | Sheet 14 | ...
    Baris 2: LINK    | (kosong) | (kosong) | (kosong) | ...
    Baris 3: KOMPONEN| tag1     | tag1     | tag1     | ...
    Baris 4+:        | tag2     | tag2     | tag2     | ...
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("❌ openpyxl tidak terinstall. Jalankan: pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "LD Mapping"

    # Sort sheet by number
    sorted_sheets = sorted(sheet_to_tags.keys())

    # ── Header styling ──
    orange_fill = PatternFill("solid", fgColor="F97316")
    yellow_fill = PatternFill("solid", fgColor="FFF3CD")
    bold_font   = Font(bold=True)
    white_font  = Font(bold=True, color="FFFFFF")

    # Baris 1: Label kolom Page
    ws.cell(row=1, column=1, value="Page").font = bold_font
    ws.cell(row=1, column=1).fill = orange_fill
    ws.cell(row=1, column=1).font = white_font

    # Baris 2: LINK GDRIVE
    ws.cell(row=2, column=1, value="LINK GDRIVE").font = bold_font
    ws.cell(row=2, column=1).fill = orange_fill
    ws.cell(row=2, column=1).font = white_font

    # Baris 3: KOMPONEN
    ws.cell(row=3, column=1, value="KOMPONEN").font = bold_font
    ws.cell(row=3, column=1).fill = orange_fill
    ws.cell(row=3, column=1).font = white_font

    # Isi setiap kolom = satu sheet
    for col_idx, sheet_num in enumerate(sorted_sheets, start=2):
        tags = sheet_to_tags[sheet_num]

        # Baris 1: nama sheet
        cell = ws.cell(row=1, column=col_idx, value=f"Sheet {sheet_num}")
        cell.font = bold_font
        cell.fill = yellow_fill
        cell.alignment = Alignment(horizontal='center')

        # Baris 2: link (kosong, diisi manual)
        ws.cell(row=2, column=col_idx, value="")

        # Baris 3+: tag number
        for row_idx, tag in enumerate(tags, start=3):
            ws.cell(row=row_idx, column=col_idx, value=tag)

    # Set lebar kolom
    ws.column_dimensions['A'].width = 15
    for col_idx in range(2, len(sorted_sheets) + 2):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 20

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"✅ File Excel disimpan: {output_path}")


# ─────────────────────────────────────────────
# STEP 3: VERIFIKASI
# ─────────────────────────────────────────────
def verify(sheet_to_tags: dict):
    sorted_sheets = sorted(sheet_to_tags.keys())
    print(f"\n📊 RINGKASAN:")
    print(f"   Total sheet unik : {len(sorted_sheets)}")
    print(f"   Nomor sheet      : {sorted_sheets[:10]}{'...' if len(sorted_sheets) > 10 else ''}")

    total = sum(len(v) for v in sheet_to_tags.values())
    print(f"   Total tag        : {total}")

    print(f"\n📋 Contoh 5 sheet pertama:")
    for sheet in sorted_sheets[:5]:
        tags = sheet_to_tags[sheet]
        print(f"   Sheet {sheet:4d} → {len(tags):3d} tag | Contoh: {tags[:3]}")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 55)
    print("📄 EXTRACT LOGIC DIAGRAM — Tag Mapping")
    print("=" * 55)

    if not os.path.exists(PDF_PATH):
        print(f"❌ File PDF tidak ditemukan: {PDF_PATH}")
        print(f"   Taruh file PDF di folder: data/")
        sys.exit(1)

    # Extract dari cross-reference index
    sheet_to_tags = extract_xref(PDF_PATH)

    if not sheet_to_tags:
        print("❌ Tidak ada data yang berhasil di-extract!")
        sys.exit(1)

    # Verifikasi
    verify(sheet_to_tags)

    # Buat Excel
    print(f"\n📝 Membuat file Excel...")
    create_excel(sheet_to_tags, OUTPUT_XLS)

    print(f"\n✅ SELESAI!")
    print(f"   1. Buka file: {OUTPUT_XLS}")
    print(f"   2. Isi kolom LINK GDRIVE (baris 2) untuk setiap sheet")
    print(f"   3. Jalankan: python update_ld.py")