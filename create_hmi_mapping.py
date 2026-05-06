"""
create_hmi_mapping.py — Buat file Excel HMI_mapping.xlsx dari hasil analisa PDF HMI
Jalankan: python create_hmi_mapping.py
Output: data/HMI_mapping.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ─────────────────────────────────────────────
# DATA HMI — hasil analisa PDF
# Format: { 'Page': [tag1, tag2, ...] }
# ─────────────────────────────────────────────
HMI_DATA = {
    'SA00': [
        'PA2704', 'PDI2704', 'FI2704', 'HC2704', 'PI1504', 'FQ1504',
        'PIC504', 'TI1504', 'PC2703', 'PI505', 'PC1505', 'HC505',
        'PDI1504B', 'PDI1504A', 'PDT1074', 'PDI1504B_RATE', 'AI1504',
        'PI2705A', 'TI2705', 'FQ2705', 'PC2705', 'FI1042B', 'FI1042BTU',
        'FQ1042B_DAILY', 'FI1042', 'FI1028',
    ],
    'SA00-1': [
        'TI1199', 'FQ1199', 'PI1199', 'HC1199', 'TI1504', 'PI1504',
        'FQ1504', 'PIC504', 'FQ1501', 'HC505', 'PC1505', 'AI1504',
        'FI1042C', 'FI1028B', 'FI640', 'FI22250', 'FI2280', 'FI1047',
        'FI1028', 'FI1042B', 'FI1042BTU', 'HC1199', 'FI640', 'B1102',
        'FI22250', 'FI2280', 'PC2703', 'PI2705A', 'TI2705', 'FQ2705',
        'PC2705', 'PDI1504B', 'PDI1504A',
    ],
    'SA01': [
        'FQI1042', 'FI1042C', 'FQI1042_DAILY', 'PC1030', 'PC1035',
        'PCA1013', 'FI1028', 'HC1003', 'FCA1015', 'TI1415', 'TI1332',
        'TA1307', 'FI1074', 'TI1029', 'PIC504', 'TI1504', 'FI1042',
        'FI1042B', 'FI1042BTU', 'SI1002A', 'SHI1028', 'XI8210', 'XI8211',
        'SI8201', 'XI8211', 'GV6201', 'FI1117', 'TA1305', 'TA1308',
        'PT1604', 'CA1002', 'PDA1155', 'PDA1156', 'LA1002',
    ],
    'SA02': [
        'PC1030', 'PC1035', 'PCA1013', 'FI1028', 'HC1003', 'FCA1003',
        'FI1004', 'HS1006', 'FIC1003', 'TI1304', 'TA1300', 'FALL1003',
        'FLLL1003', 'FCA1056', 'FI1044', 'HC1021', 'PI1044B', 'FI1871',
        'FI1872', 'TI1831', 'TI1832', 'TI1833', 'TI1834', 'FCA156',
        'TCA1011', 'HZ1023', 'HC1023', 'PZ1004', 'GV6101', 'ECV6101',
        'SI1001A', 'CA1001', 'PC6106', 'PI1870', 'FI1879', 'PDAH101',
        'KPH101J', 'SI1001A', 'SI6101', 'PI9102', 'PI9103', 'PI9104',
        'PI9105', 'PI9106', 'PI9107', 'TI9101', 'TI9102', 'TI9103',
        'TI9104', 'TI9105', 'TI9106', 'TI9107',
    ],
    'SA03': [
        'TA1309', 'FI1028', 'HC1003', 'PC1002', 'PC1003', 'FI1047',
        'PI1221A', 'PI1221B', 'PI1221C', 'PAL1120', 'PAH1121', 'PA1228',
        'FLLL1201', 'FFY1201', 'FFAL1201', 'TA1313', 'FI1047',
        'LAH1150', 'LAH1160', 'SI101-BJ2', 'PCA1001', 'HC1039',
        'HC1038', 'HC1037', 'HC1036', 'HC1035', 'PI1061', 'PI1062',
        'PI1063', 'PI1064', 'PI1065', 'PSL1161', 'PSL1162', 'PSL1163',
        'PSL1164', 'PSL1165', 'PA1001',
    ],
    'SA04': [
        'HC1021', 'HY1021B', 'TCA1305', 'TA1419', 'TI1415', 'TI1332',
        'AA1010A', 'AA1010B', 'TI1307', 'TI1326', 'TI1327', 'TI1331',
        'TI1332', 'TA1312', 'TI1313', 'TC1311', 'PA1057', 'PA1057A',
        'TI1417', 'TI1418', 'TA1308', 'TA1309', 'PDA1155', 'PDA1156',
        'SI101-BJ2', 'SI101-BJ1', 'PZ1019', 'LCA1855', 'PCA1855',
        'FI1810', 'PI1809', 'PDH1063', 'PDH1067C', 'TI1334', 'BCS',
        'AA1011', 'AI1001',
    ],
    'SA05': [
        'TA1312', 'TA1314', 'TA1051', 'TA1052', 'TA1053', 'TA1062',
        'LAL1051', 'LAL1053', 'LAL1054', 'LAL1055', 'LC1161', 'LC1142',
        'LC1143', 'LC1144', 'LC1145', 'LC1141', 'LC1142', 'FI1141',
        'FI1142', 'FI1143', 'FI1144', 'FI1145', 'FA1094', 'FA1151',
        'PC1113', 'PV1018A', 'PV1018B', 'PCA1018', 'AI1005', 'TC1010',
        'TI1334', 'TI1334B', 'TI1340', 'TI1335', 'TI1338', 'TI1339',
        'TI1337', 'TI1550', 'PT1618', 'PCA1018', 'AI1005',
    ],
    'SA06': [
        'FI1201', 'FCA1001', 'FCA1002', 'HC1002', 'N2', 'FI1201',
        'LCA1003', 'LCA1005', 'TA1430', 'TI1302', 'TI1362', 'TI1310',
        'TI1024', 'FFLL1201', 'FFAL1201', 'LX1900B', 'LX1900C',
        'PA1053', 'PA1053C', 'FI1202', 'LCA1025', 'TI1405', 'AA1017',
        'AI1007', 'PC2703', 'PDI1069', 'HC6401', 'HC6402',
    ],
    'SA07': [
        'MOV1005', 'TCA1011', 'TI1341', 'TI1342', 'TI1343', 'TI1344',
        'TI1346', 'TI1347', 'TI1348', 'TI1349', 'TI1346', 'TI1779',
        'PDA1063', 'AI1020', 'PDA1111', 'PI1614', 'PC1030', 'HS1008',
        'HS1009', 'TI1613', 'HC1009', 'HC1007', 'MOV1007',
        'TA1339', 'FI1103', 'TI1411', 'TI1412', 'TCA1011',
    ],
    'SA09': [
        'FCA1016', 'TC1006', 'PCA1104', 'FI1090', 'PDA1043', 'TI1406',
        'TI1654', 'LCA1041', 'LALL1106', 'FC1009', 'LA1017', 'FC1050',
        'FC1017', 'LCA1042', 'LCA1020', 'LALL1103', 'TI1407', 'TI1409',
        'TI1410', 'FI1040', 'FCA1040', 'PCA1104', 'PC1833', 'LCA1162',
        'FI1049', 'PX10338', 'PI1049',
    ],
    'SA10': [
        'TI1412', 'FI1040', 'TI1420', 'TI1421', 'TI1362', 'LALL1103',
        'LCA1020', 'FCA1003', 'FC1050', 'LCA1003', 'LALL1205', 'AI1003',
        'TI1430', 'HS1018', 'HS1019', 'HS1018RS', 'TI1361', 'LX1900B',
        'LX1900C', 'FI1049',
    ],
    'SA11': [
        'AI1030', 'AI1031', 'TA1200', 'TA1201', 'TA1202', 'TA1357',
        'TA1358', 'TA1360', 'TA1361', 'TA1362', 'TA1363', 'TA1365',
        'TA1366', 'TCA1012', 'PDI1004', 'LCA1068', 'TI1363', 'TI1616',
        'TI1618', 'PC1008', 'LAH1132', 'TI1353', 'PDA1126', 'LC1132',
        'LCA1039', 'V1211', 'TI1250', 'PI117', 'PCA1004', 'PC1006',
        'HS1011', 'TA1359', 'TA1369', 'TA1381',
    ],
    'SA12': [
        'SI1003A', 'SI1003A', 'GV8301', 'ECV8301', 'TI1750', 'HC1041',
        'KPH101J', 'PI1045', 'FI1628', 'TA1367', 'LCA1011', 'LCA1159',
        'FI1074', 'FI1109', 'LAH1111', 'LAH1112', 'PC1114', 'TA1364',
        'TA1792', 'FI1075', 'TCA1012', 'PI1364', 'PDA1044', 'FC1007',
        'FC1008', 'TDI116C', 'TDI116C', 'TI6315', 'PI63015',
    ],
    'SA13': [
        'HC1026', 'HC1042', 'TI1373', 'TI1374', 'AI1021A', 'AI1021B',
        'AI1021C', 'AI1021D', 'AI1021E', 'AF1021D', 'TA1369', 'TA1371',
        'FC1020', 'FC1077', 'LAHH1218', 'LALL1218', 'LA1218',
        'LC1013', 'SD105J', 'PI2SI', 'TC1030', 'TI1600', 'TI1606',
        'TI1633', 'PDI1121', 'PDA1059', 'HS1020', 'HS1621', 'MOV1004',
    ],
    'SA14': [
        'TI1374', 'TI1375', 'TI1377', 'TI1378', 'TI1379', 'TI1380',
        'TI1381', 'TI1382', 'TI1383', 'TI1384', 'TI1385', 'TI1386',
        'TI1387', 'TI1388', 'TI1389', 'TI1390', 'TI1391', 'TI1392',
        'TI1394', 'AI1040', 'AI1041', 'AI1042', 'AI1043',
        'XA63115A', 'XA63115B', 'XA63119', 'XA6312A', 'XA6312B',
        'XA63120', 'TI1373', 'PDI1054', 'LC1013', 'PC1029', 'FC1029',
        'HC1025', 'MOV1004', 'TA1397', 'PA1051', 'HC1060',
        'FI1267', 'FI1257', 'PIB1257',
    ],
    'SA15': [
        'XV1171', 'XV1170', 'MV1917', 'TI1044', 'TI1043', 'TI1042',
        'TI1041', 'TA1368', 'TC1040', 'TAHH1040', 'ZLC1040A', 'ZLC1040B',
        'PC1047', 'PC1048', 'PDI1716', 'LCA1166', 'PCA1131', 'PDI1628A',
        'PDI1628B', 'PDI1712', 'FI1046', 'PI1628', 'AI1014', 'AI1016',
        'XY1192', 'XY1190', 'MV1915', 'MV1916', 'XY1191', 'XV1160',
        'XV1164', 'XV1165',
    ],
    'SA16': [
        'FI1058E', 'FC1078', 'LCA1012', 'PCA1108', 'PCA1109',
        'PAHH1070', 'TI1403', 'FI1058', 'FI1058_JAM', 'FQI1066C',
        'FQI1066C', 'FAL1060', 'FI1560', 'PC1028', 'PC1029',
        'LCA1015', 'LCA1024', 'TC1796', 'TI1402', 'FC1079', 'FC1078',
        'LY1012A', 'LY1012B', 'LAL1113', 'LAL1121', 'LAL1122',
        'LALL1113', 'PAH6630', 'FAL1061', 'FI1061', 'FA1060',
        'FQI1061C', 'FA1061C', 'FI1061C', 'XY1074', 'NH3_RATE',
        'NH3_IMPOR', 'PIC1029', 'FIC1029',
    ],
    'SA17': [
        'ZI6512B', 'ZI6511B', 'TI6510', 'TI6517', 'PCA1020', 'GV6501',
        'FI1037C', 'FQI1037', 'FC1037', 'PC1037', 'TI1401', 'TI1037',
        'PI1037A', 'FI1109', 'FI1060', 'TC1796', 'TI1402', 'LCA1012',
        'FC1009', 'FC1010', 'FC1011', 'FC1012', 'LCA1021', 'LCA1022',
        'LCA1023', 'LCA1024', 'LSHH1214', 'LSHL1215', 'LSHL1216',
        'LSHL1217', 'TA1398', 'TA1399', 'TA1400', 'TA1404', 'TAHH1404',
        'FI1061', 'FAL1061', 'SI1005A', 'SI6601', 'ACV6501',
        'PDI1121', 'PC501', 'FI1278', 'FI1029',
    ],
    'SA17A': [
        'FC1037', 'FI1037C', 'PC1037', 'PI1037A', 'TI1037',
        'FCS198', 'FCS199', 'PC5198', 'PC5199', 'FQ5199', 'FI5198',
        'TI5198', 'PI5198', 'HCS198', 'HCS196',
    ],
    'SA18': [
        'FT1006', 'FC1006', 'FI1033B', 'FFY1033', 'FCA1033', 'FC1001',
        'LC1001', 'LC1101', 'LI1001A', 'LI1001B', 'LALL1060', 'LAH1029',
        'LAHH101F', 'LT1060', 'LI1059', 'LI1060', 'LALL1223', 'LT1001B',
        'LC1001B', 'TI1411', 'TI1412', 'TCA1011', 'FI1043', 'HC6401',
        'FYY10600', 'FI11060D', 'SI6401', 'SI1004A', 'FA1108',
        'FALL1106', 'FLLL1106', 'PI1034', 'TI1370', 'TCA1011',
        'LCA1129', 'FC1020', 'PI1704', 'TI1558', 'TI1659', 'TI1658',
    ],
    'SA19': [
        'HS1102', 'TI1411', 'FC1006', 'AI1015', 'TI1370', 'LV1001A',
        'LV1001B', 'LSL1125', 'LAH1124', 'LALL1126', 'PI1704', 'AI1007',
        'LCA1030', 'FCA1097', 'PI1087', 'HS1102', 'FA1108', 'FALL1106',
        'FLLL1106', 'GV6401', 'HC5401', 'CA1004', 'CA1004A',
        'PSLL6402', 'SI1004A', 'FYY10900', 'FI11090D',
    ],
    'SA20': [
        'FI1031', 'TCA1005_RATE', 'PI1788', 'PCA1018', 'FC1006_OPNUM',
        'PAH1141', 'PAL1140', 'TCA1005', 'MZ1033', 'MOV1033', 'HC1028',
        'HC1029', 'HV1028', 'TI1753', 'TI1750', 'TI1751', 'KPH101J',
        'TA1896', 'FI1115', 'HC5401', 'PDC1022', 'TAHH1005', 'TI1553',
        'PC1039', 'FI1071', 'FI1091', 'FI1091B', 'SP154', 'SP170',
        'TCA1022', 'TCA1020', 'TI1554', 'GV6101', 'ECV6101', 'GV6301',
        'ECV6301', 'PY1018A', 'PY1018B', 'FCA1002', 'FI1402', 'FI1092',
        'HC1402', 'PC1012', 'PC1035', 'PCA1013', 'FQ1402',
    ],
    'SA21': [
        'PI1071', 'TI1560', 'FI1091', 'FI1091B', 'FQ1091', 'TCA1022',
        'HC1402', 'FI1092', 'PC1012', 'PC1016', 'PC1017', 'PC1018',
        'PC1019', 'PC1020', 'PCA1017', 'PCA1009', 'TC1021', 'SP138',
        'SP153', 'TI1555', 'GV8301', 'ECV8301', 'GV6101', 'ECV6101',
        'GV6201', 'ECV6201', 'ACV8301', 'SI1005A', 'SI1002A', 'FI2205',
        'TA1811', 'FI1116', 'FI1278', 'FI1265', 'TI1754',
    ],
    'SA22': [
        'TI1757', 'PA1068', 'LCA1019', 'LSHH1128', 'FC1032', 'TI1561',
        'AA1018', 'FA1094', 'TI1119', 'TI1600', 'TI1645',
        'DELTA_TEMP_127C', 'FI1600', 'TI1611', 'TI1612',
    ],
    'SA23': [
        'FI1037C', 'FQI1037', 'PC1037', 'FC1037', 'FCS198', 'PC5198',
        'PC5199A', 'FCS199', 'PC5199', 'FI5199', 'FQ5199',
        'PI5198', 'TI5198', 'LCA1022', 'LCA1021', 'LAHH1214',
        'CF3', 'CF4', 'FI1064', 'TAl1389', 'TI1363', 'LCA1015',
        'LY1012B', 'FC1078', 'FC1079', 'FI1066C', 'FI1064',
        'FAL1061', 'FI1061',
    ],
    'SA24': [
        'FRC1011', 'FA1012', 'FI1075', 'PRC1037', 'PRC10211', 'TSH1018',
        'TCA1017', 'PC1038', 'FC1038', 'PA1038A', 'PAHH1038', 'FI1038',
        'TI1038', 'XV1038', 'PDR1026', 'FR1029', 'PDR1031', 'FR1035',
        'FI1035',
    ],
    'SA25': [
        'PDAH1115', 'LA6603', 'LI1119',
    ],
    'SA25A': [
        'FCA1001', 'LCA1030', 'LCA1035', 'LC1001', 'FCA1004',
        'PC1005', 'PC1006', 'PCA1010', 'TCA1005', 'TCA1065',
        'TCA1001', 'LC1013', 'LCA1024', 'PCA1013', 'PCA1014',
        'PCA1017', 'HC1043', 'FI1119', 'FI1277', 'FI1115',
        'NH_RATIO', 'H2N2_RATIO', 'RATE_NH3', 'TA1336',
        'PC1033', 'PI1045', 'FI1092', 'FI1095', 'PI1504',
        'FI1058E', 'PCA1013', 'FI1037C',
    ],
    'SA25B': [
        'LC1001', 'LC1001B', 'FC1033', 'FC1006', 'LCA1030', 'LCA1035',
        'TCA1005', 'TCA1011', 'TCA1013', 'TCA1014', 'TCA1015',
        'TCA1016', 'TCA1017', 'TCA1020', 'TCA1022', 'PCA1013',
        'PCA1015', 'PCA1017', 'HC1045',
    ],
    'SA26': [
        'LC6305', 'LC6304', 'LI6301', 'LI6303', 'LIC6301', 'LI6303',
        'LIC6201', 'LI6303', 'LI6303', 'PI6302', 'PI11045',
        'LV6302B', 'LI6004', 'PI6005', 'PI6111', 'PI6211', 'PI6311',
        'PI6511', 'PC6123', 'PC6123', 'PI6309', 'PDC6601', 'PDC6602',
        'SI6101', 'SI6201', 'SI8101', 'SI8201', 'SI6601',
        '102-JL-J1', '102-JL-J2', '102-JL-J3', '102-JLJ1A',
        '102-JLJ1M', '102-JLJ2', '102-JLF1',
    ],
    'SA27': [
        'XV1251', 'XV1250', 'XV1256', 'XV1255', 'HC1050',
        'PSHH1251', 'PSLL1250', 'PAH1151', 'PAL1150', 'PAL1158',
        'PSL1157', 'PI1784',
    ],
    'SA27A': [
        'PIL1484', 'PIL1160', 'PIL1762', 'PIL1764', 'PIL1765',
        'PIL1766', 'PIL1769', 'PILI170', 'PIL1763',
        'XV1220', 'XYV1259A', 'XYV1259B', 'XYV1220',
        'HORN', 'ROTATING_LIGHT',
    ],
    'SA28': [
        'MOV1017', 'MOV1015', 'XV1160', 'XV1164', 'PV1047',
        'MOV1013', 'MOV1016', 'XV1161', 'XV1165', 'PV1048',
        'XV1162', 'PC1049', 'PC1047', 'PC1048', 'PC1131',
        'TI1041', 'TI1043', 'TC1040', 'PCA1131', 'KIC1001',
        'HS1912', 'PRCSRDY', 'ZLC1040A', 'ZLC1040B',
    ],
    'SA28A': [
        'PT1047', 'PT1048', 'PC1131', 'TT1040', 'TT1041', 'TT1043',
    ],
    'SA29': [
        'PC1002', 'FI1028', 'HC1003', 'XV1220', 'XV1221',
        'PSHH1221', 'PSLL1220', 'HC1036', 'HC1037', 'HC1038', 'HC1039',
        'PI1061', 'PI1062', 'PI1063', 'PI1064', 'PI1065',
        'PSL1161', 'PSL1162', 'PSL1163', 'PSL1164', 'PSL1165',
        'PAH1221', 'PAL1220', 'LAH1161', 'PI1784',
    ],
    'SA29A': [
        'PIL1484', 'PIL1162', 'PIL1163', 'PIL1164', 'PIL1165',
        'PIL1166', 'PIL1168', 'PIL1769', 'PIL1763',
        'XV1220', 'XYV1222', 'XYV1220', 'XY1220_MAIN_FUEL',
    ],
    'SA30': [
        'TCA1005', 'TA1741', 'TA1742', 'TA1743', 'TA1744',
        'TA1745', 'TA1746', 'TA1788', 'XV1240', 'XV1241',
        'PAH1146', 'PAL1140', 'PSHH1241', 'PSLL1240', 'PSH1241',
        'PSL1240', 'LAH1160', 'XV1245', 'XV1246', 'PI1788',
        'PAH1146', 'PAL1145', 'PSL1147', 'PSHH1246',
        'PDC1022', 'FI1947', 'LAH1150',
    ],
    'SA30A': [
        'PIL1484', 'PIL1140', 'PIL1141', 'PIL1142', 'PIL1143',
        'PIL1144', 'XV1240', 'XV1241', 'XY1240', 'XY1241',
    ],
    'SA31': [
        'FCA1001', 'FCA1002', 'FFIC1003', 'FIC1003', 'LAHH1218',
        'LALL1218', 'XS2012', 'PAHH1059', 'PSL1161-S', 'FLLL1201',
        'FAL1152', 'PCA1018_RAMP', 'FAL1003', 'IL6101_PERMIT',
        'COMMON_SHUTDOWN_101-J_COMMSHDN', 'MOV1009', 'FAL1061',
        'LAHH191F', 'ZSC1004', 'ZSO1004', 'ZSC1006', 'ZSO1006',
    ],
    'SA32': [
        'PSLL6102', 'ZS6101', 'SAH6101', 'SSH1001', 'ZS6102', 'ZS6103',
        'PSL6000', 'PDSH6003', 'LSH6006', 'LSH6000', 'LSH6004',
        'PDSL6104', 'LSH6102', 'LSH6103', 'LSH6104', 'PSL6104',
        'HS6107', 'HS6106',
    ],
    'SA32C': [
        'ZI6110A', 'ZI6110B', 'ZI6111A', 'ZI6111B', 'ZI6112A', 'ZI6112B',
        'XI6111A', 'XI6111B', 'XI6112A', 'XI6112B', 'XI6113A', 'XI6113B',
        'XI6114A', 'XI6114B', 'XI6115A', 'XI6115B', 'XI6116A', 'XI6116B',
        'TE6101', 'TE6102', 'TE6103', 'TE6104', 'TE6105', 'TE6109',
        'TE6110', 'TE6111', 'TE6112', 'TE6113', 'TE6114', 'TE6115',
        'TE6116', 'TE6117', 'TE6118', 'KPH101J',
    ],
    'SA33': [
        'FI1037C', 'FQ1037', 'FC1037', 'PC1037', 'TA1401', 'TI1037',
        'PI1037A', 'FCS198', 'FCS199', 'PC5198', 'PC5199', 'PC5199A',
        'LCA1021', 'LCA1022', 'LCA1023', 'LCA1024', 'LAHH1214',
        'CF3', 'CF4', 'LY1012B', 'FC1078', 'FC1079', 'FI5198',
        'FI5199', 'SI1005A', 'PCA1020', 'XY6502', 'HC1030',
        'TA1398', 'TA1399', 'TA1400', 'TAHH1404', 'FI1061',
    ],
    'SA33A': [
        'FI1037C', 'FQ5199', 'FC1037', 'PC1037', 'TA1401', 'TI1037',
        'PI1037A', 'FCS198', 'FCS199', 'PC5198', 'PC5199', 'HCS198',
        'MOV1101', 'MOV1102', 'HV1002', 'LCA1021', 'LCA1022',
        'FI1064', 'TI1363', 'LCA1039', 'FC1078', 'FC1079', 'LY1012B',
        'FI5198', 'FI5199', 'SI1005A',
    ],
    'SA34': [
        'PSH6000', 'PDSH6003', 'LSH6006', 'TSH6000', 'LSH6002',
        'PDSL6104', 'LSH6101', 'LSH6102', 'LSH6103', 'PSL6204',
        'LSL6202', 'PSLL6202', 'DCS_XA6201', 'XA6201', 'HS1202',
        'HS6201_ACK', 'PB102J', 'HS6202_RST', 'HS6206_EMERG_PB',
        'IL5201', 'HN537-102',
    ],
    'SA34C': [
        'ZI6210A', 'ZI6210B', 'ZI6211A', 'ZI6211B',
        'XI6211A', 'XI6211B', 'XI6212A', 'XI6212B', 'XI6213A', 'XI6213B',
        'XI6214A', 'XI6214B', 'TI6201', 'TI6202', 'TI6203',
        'TI6206', 'TI6208', 'TI6209', 'TI6210', 'TI6211',
        'TI6213', 'TI6215', 'XI6210',
    ],
    'SA35': [
        'SSH1003', 'ZS6302', 'TSH6301', 'TSH6309', 'TSH6317',
        'XSH6311', 'XSH6312', 'XSHH6312', 'LSH6301', 'LSH6303',
        'LSH6305', 'PDSL6304', 'SSH6302', 'LSL6302', 'LSLL6302',
        'DCS_XA6301', 'XA6310', 'HS1203', 'HS6301_ACK', 'PB103J',
        'HS6302_RST', 'HS6306_EMERG_PB', 'IL6301-OCS',
        'HN537-1', 'X6301_3', 'X6301_2',
    ],
    'SA35C': [
        'ZI6310A', 'ZI6310B', 'ZI6311A', 'ZI6311B', 'ZI6312A', 'ZI6312B',
        'XI6311A', 'XI6311B', 'XI6312A', 'XI6312B', 'XI6313A', 'XI6313B',
        'XI6314A', 'XI6314B', 'XI6315A', 'XI6315B', 'XI6316A', 'XI6316B',
        'TI6301', 'TI6302', 'TI6303', 'TI6305', 'TI6306', 'TI6309',
        'TI6310', 'TI6311', 'TI6315', 'TI6316', 'TI6317', 'TI6318',
        'TI6319', 'TI6323', 'TI6324', 'XI6310',
    ],
    'SA36': [
        'SSH1005', 'TSH6501', 'ZS6502', 'PDSL6505', 'PDSL6507',
        'XSH6511', 'XSHH6511', 'XSH6512', 'XSHH6512', 'ZSH6510',
        'LSL6501', 'PSLL6502_LOLO', 'PSLL6504_LOLO', 'SSH8H6502',
        'SSA601_105', 'DCS_XA6501', 'XA6510', 'HS1205',
        'HS6501_ACK', 'PB105J', 'HS6502_RST', 'HS6506_EMERG_PB',
        'IL6501', 'HN537-1', 'X6501_3', 'X1205_2',
    ],
    'SA36C': [
        'ZI6510A', 'ZI6510B', 'ZI6511A', 'ZI6511B', 'ZI6512A', 'ZI6512B',
        'XI6511A', 'XI6511B', 'XI6512A', 'XI6512B', 'XI6513A', 'XI6513B',
        'XI6514A', 'XI6514B', 'XI6515A', 'XI6515B', 'XI6516A', 'XI6516B',
        'TI6501', 'TI6502', 'TI6503', 'TI6505', 'TI6506', 'TI6507',
        'TI6509', 'TI6510', 'TI6511', 'TI6513', 'TI6514', 'TI6515',
        'TI6517', 'TI6518', 'TI6519', 'TI6521', 'TI6522', 'TI6523',
        'XI6510',
    ],
    'SA37': [
        'FI1042B', 'FI1028', 'FI1092', 'FI1058', 'HV1402', 'FI1091',
        'FQA1501', 'FQA2701', 'FQA2704', 'GHV', 'PERTAGAS',
        'PHE_WMO', 'LISTRIK',
    ],
    'SA38': [
        'XV1501A', 'XV1501B', 'XV1501C', 'AI1040', 'AI1501',
        'AI1049', 'AI1500',
    ],
    'SA39': [
        'FC1033', 'FI1277', 'FI1115', 'FI1091', 'FI1091B',
        'FQI1091B', 'FI1092', 'PCA1002', 'FI1103', 'FI1275',
        'FI1276', 'FC1050', 'FI1049', 'FI2704', 'FQ1501',
        'FI2701', 'FI2280', 'FI22250', 'FI640',
    ],
}

def create_excel():
    os.makedirs("data", exist_ok=True)
    OUTPUT = "data/HMI_mapping.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "HMI Mapping"

    orange_fill = PatternFill("solid", fgColor="F97316")
    yellow_fill = PatternFill("solid", fgColor="FFF3CD")
    white_font  = Font(bold=True, color="FFFFFF")
    bold_font   = Font(bold=True)

    pages = list(HMI_DATA.keys())

    # Baris 1: Page
    ws.cell(row=1, column=1, value="Page").fill = orange_fill
    ws.cell(row=1, column=1).font = white_font

    # Baris 2: LINK GDRIVE
    ws.cell(row=2, column=1, value="LINK GDRIVE").fill = orange_fill
    ws.cell(row=2, column=1).font = white_font

    # Baris 3: KOMPONEN
    ws.cell(row=3, column=1, value="KOMPONEN").fill = orange_fill
    ws.cell(row=3, column=1).font = white_font

    for col_idx, page in enumerate(pages, start=2):
        tags = HMI_DATA[page]

        cell = ws.cell(row=1, column=col_idx, value=page)
        cell.font = bold_font
        cell.fill = yellow_fill
        cell.alignment = Alignment(horizontal='center')

        ws.cell(row=2, column=col_idx, value="")  # link diisi manual

        for row_idx, tag in enumerate(tags, start=3):
            ws.cell(row=row_idx, column=col_idx, value=tag)

    ws.column_dimensions['A'].width = 15
    for col_idx in range(2, len(pages) + 2):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 18

    wb.save(OUTPUT)
    print(f"✅ File Excel dibuat: {OUTPUT}")

    total_tags = sum(len(v) for v in HMI_DATA.values())
    print(f"📊 Total page  : {len(pages)}")
    print(f"📊 Total tag   : {total_tags}")
    print(f"\n📋 Daftar page:")
    for page, tags in HMI_DATA.items():
        print(f"   {page:10s} → {len(tags):3d} tag")

if __name__ == "__main__":
    print("=" * 50)
    print("📄 BUAT HMI MAPPING EXCEL")
    print("=" * 50)
    create_excel()
    print(f"\n✅ Selesai!")
    print(f"   1. Buka data/HMI_mapping.xlsx")
    print(f"   2. Isi link GDrive di baris 2 per page")
    print(f"   3. Jalankan: python update_hmi_v2.py")