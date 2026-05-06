"""
update_ld.py — Update kolom Logic_Diagram di database dari file LD_mapping.xlsx

Cara pakai:
    python update_ld.py

Pastikan sebelumnya:
    1. Sudah jalankan extract_ld.py → menghasilkan data/LD_mapping.xlsx
    2. Sudah isi kolom LINK GDRIVE (baris 2) di LD_mapping.xlsx
    3. Kolom Logic_Diagram sudah ada di database:
       ALTER TABLE equipment ADD COLUMN Logic_Diagram TEXT DEFAULT '-';
"""

import sqlite3
import os
import re
from openpyxl import load_workbook

DB_PATH    = "data/ammonia.db"
EXCEL_PATH = "data/LD_mapping.xlsx"

def normalize_tag(tag: str) -> str:
    """Normalisasi tag: hapus spasi, strip, underscore → uppercase."""
    return re.sub(r'[\s\-_]', '', str(tag)).upper()

def read_ld_excel() -> dict:
    """
    Baca LD_mapping.xlsx dan return:
    { normalized_tag: (original_tag, hmi_link) }
    """
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ File tidak ditemukan: {EXCEL_PATH}")
        print(f"   Jalankan dulu: python extract_ld.py")
        return {}

    print(f"📖 Membaca file: {EXCEL_PATH}")
    wb   = load_workbook(EXCEL_PATH, read_only=True)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Baris 1: Page header (index 0)
    # Baris 2: LINK GDRIVE (index 1)
    # Baris 3+: Tag number (index 2+)

    if len(rows) < 3:
        print("❌ Format Excel tidak sesuai.")
        return {}

    link_row = rows[1]   # index 1 = baris 2 = LINK GDRIVE
    tag_mapping = {}

    skipped_no_link = 0

    for col_idx in range(1, len(link_row)):
        ld_link = link_row[col_idx] if col_idx < len(link_row) else None

        # Skip kolom yang tidak punya link
        if not ld_link or str(ld_link).strip() in ['-', '', 'None']:
            skipped_no_link += 1
            continue

        ld_link = str(ld_link).strip()

        # Ambil semua tag di kolom ini (baris 3 ke bawah)
        for row_idx in range(2, len(rows)):
            cell = rows[row_idx][col_idx] if col_idx < len(rows[row_idx]) else None
            if not cell or str(cell).strip() in ['', 'None']:
                continue

            tag_orig = str(cell).strip()
            tag_norm = normalize_tag(tag_orig)

            if len(tag_norm) < 3:
                continue

            tag_mapping[tag_norm] = (tag_orig, ld_link)

    wb.close()

    if skipped_no_link > 0:
        print(f"ℹ️  {skipped_no_link} sheet dilewati (belum ada link)")

    print(f"✅ {len(tag_mapping)} tag siap diupdate")
    return tag_mapping


def build_db_index(cursor) -> dict:
    """Buat index semua tag di database {normalized: original}."""
    cursor.execute("SELECT tag_number FROM equipment")
    return {normalize_tag(r[0]): r[0] for r in cursor.fetchall()}


def update_ld_links(tag_mapping: dict):
    if not tag_mapping:
        print("⚠️  Tidak ada data untuk diupdate.")
        return

    if not os.path.exists(DB_PATH):
        print(f"❌ Database tidak ditemukan: {DB_PATH}")
        return

    # Cek apakah kolom Logic_Diagram ada
    conn   = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    cursor.execute("PRAGMA table_info(equipment)")
    columns = [row[1] for row in cursor.fetchall()]
    if 'Logic_Diagram' not in columns:
        print("❌ Kolom 'Logic_Diagram' tidak ada di database!")
        print("   Jalankan di DB Browser:")
        print("   ALTER TABLE equipment ADD COLUMN Logic_Diagram TEXT DEFAULT '-';")
        conn.close()
        return

    # Buat index database
    db_index = build_db_index(cursor)
    print(f"📊 Total tag di database: {len(db_index)}")

    updated   = 0
    notfound  = 0
    notfound_list = []

    for norm_tag, (orig_tag, ld_link) in tag_mapping.items():
        if norm_tag in db_index:
            db_tag = db_index[norm_tag]
            cursor.execute('''
                UPDATE equipment SET Logic_Diagram = ? WHERE tag_number = ?
            ''', (ld_link, db_tag))
            updated += 1
        else:
            notfound += 1
            notfound_list.append(orig_tag)

    conn.commit()
    conn.close()

    print(f"\n📊 HASIL UPDATE:")
    print(f"   ✅ Berhasil diupdate : {updated} tag")
    print(f"   ❌ Tidak ditemukan   : {notfound} tag")

    if notfound_list:
        print(f"\n⚠️  Contoh tag tidak ditemukan (10 pertama):")
        for t in notfound_list[:10]:
            print(f"   - '{t}'")


def verify():
    if not os.path.exists(DB_PATH):
        return
    conn   = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    cursor.execute('''
        SELECT COUNT(*) FROM equipment
        WHERE Logic_Diagram != '-' AND Logic_Diagram != '' AND Logic_Diagram IS NOT NULL
    ''')
    count = cursor.fetchone()[0]
    print(f"\n🔗 Total equipment dengan Logic Diagram link: {count}")

    print("\n📋 Contoh 5 data:")
    cursor.execute('''
        SELECT tag_number, Logic_Diagram FROM equipment
        WHERE Logic_Diagram != '-' AND Logic_Diagram != ''
        LIMIT 5
    ''')
    for r in cursor.fetchall():
        print(f"   {r[0]:15s} → {str(r[1])[:55]}...")

    conn.close()


if __name__ == "__main__":
    print("=" * 55)
    print("🔗 UPDATE LOGIC DIAGRAM LINK — Ammonia Plant")
    print("=" * 55)

    tag_mapping = read_ld_excel()

    if tag_mapping:
        print(f"\n📋 Preview 5 mapping pertama:")
        for i, (norm, (orig, link)) in enumerate(list(tag_mapping.items())[:5]):
            print(f"   {orig:20s} → {link[:45]}...")

        print(f"\nTotal tag akan diupdate: {len(tag_mapping)}")
        confirm = input("\nLanjutkan update database? (y/n): ")

        if confirm.lower() == 'y':
            update_ld_links(tag_mapping)
            verify()
            print("\n✅ Selesai! Restart backend: python backend.py")
        else:
            print("❌ Update dibatalkan.")
    else:
        print("\n❌ Tidak ada data. Pastikan LD_mapping.xlsx sudah diisi linknya.")